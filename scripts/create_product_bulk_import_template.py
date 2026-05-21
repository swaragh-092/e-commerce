from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUT = Path("docs/product_bulk_import_template.xlsx")


SHEETS = {
    "Products": {
        "headers": [
            "product_key",
            "name",
            "slug",
            "sku",
            "type",
            "status",
            "is_enabled",
            "is_featured",
            "brand_id",
            "brand_slug",
            "price",
            "sale_price",
            "sale_start_at",
            "sale_end_at",
            "sale_label",
            "quantity",
            "weight_grams",
            "length_cm",
            "breadth_cm",
            "height_cm",
            "short_description",
            "description",
            "meta_title",
            "meta_description",
            "meta_keywords",
            "og_image",
            "tax_config_json",
        ],
        "rows": [
            [
                "P001",
                "Cotton T Shirt",
                "cotton-t-shirt",
                "TSHIRT-001",
                "simple",
                "published",
                "true",
                "false",
                "",
                "nike",
                499,
                399,
                "2026-05-20T00:00:00Z",
                "2026-06-01T00:00:00Z",
                "Summer Sale",
                100,
                500,
                10,
                10,
                3,
                "Soft cotton tshirt",
                "<p>Full product description</p>",
                "Cotton T Shirt",
                "Buy cotton tshirt online",
                "tshirt,cotton",
                "https://site.com/og.jpg",
                '{"isCustom":true,"inclusive":true,"sgst":0.09,"cgst":0.09,"igst":0.18}',
            ],
            [
                "P002",
                "Variable T Shirt",
                "variable-t-shirt",
                "",
                "variable",
                "draft",
                "true",
                "false",
                "",
                "nike",
                499,
                "",
                "",
                "",
                "",
                "",
                500,
                10,
                10,
                3,
                "Tshirt with size and color variants",
                "<p>Description</p>",
                "",
                "",
                "",
                "",
                "",
            ],
        ],
    },
    "ProductCategories": {
        "headers": ["product_key", "category_id", "category_slug", "sort_order"],
        "rows": [["P001", "", "mens-clothing", 0], ["P001", "", "tshirts", 1]],
    },
    "ProductImages": {
        "headers": ["product_key", "variant_key", "image_url", "media_id", "alt", "sort_order", "is_primary"],
        "rows": [
            ["P001", "", "https://site.com/images/tshirt-main.jpg", "", "Cotton T Shirt Main", 0, "true"],
            ["P002", "P002-V001", "https://site.com/images/red-small.jpg", "", "Red Small Variant", 0, "true"],
        ],
    },
    "ProductTags": {
        "headers": ["product_key", "tag_name"],
        "rows": [["P001", "new arrival"], ["P001", "cotton"], ["P001", "summer"]],
    },
    "ProductAttributes": {
        "headers": [
            "product_key",
            "attribute_id",
            "attribute_slug",
            "value_id",
            "value_slug",
            "custom_name",
            "custom_value",
            "is_variant_attr",
            "sort_order",
        ],
        "rows": [
            ["P002", "", "color", "", "red", "", "", "true", 0],
            ["P002", "", "size", "", "small", "", "", "true", 1],
            ["P001", "", "", "", "", "Material", "100% Cotton", "false", 0],
        ],
    },
    "ProductVariants": {
        "headers": ["product_key", "variant_key", "sku", "price", "stock_qty", "reserved_qty", "is_active", "sort_order", "media_id"],
        "rows": [
            ["P002", "P002-V001", "TSHIRT-RED-S", 499, 25, 0, "true", 0, ""],
            ["P002", "P002-V002", "TSHIRT-RED-M", 499, 30, 0, "true", 1, ""],
        ],
    },
    "VariantOptions": {
        "headers": ["variant_key", "attribute_id", "attribute_slug", "value_id", "value_slug"],
        "rows": [
            ["P002-V001", "", "color", "", "red"],
            ["P002-V001", "", "size", "", "small"],
            ["P002-V002", "", "color", "", "red"],
            ["P002-V002", "", "size", "", "medium"],
        ],
    },
    "ProductTabs": {
        "headers": ["product_key", "title", "content", "type", "sort_order", "is_active"],
        "rows": [
            ["P001", "Shipping Info", "<p>Ships within 2 days.</p>", "html", 0, "true"],
            ["P001", "Care Instructions", "<p>Machine wash cold.</p>", "html", 1, "true"],
        ],
    },
    "ComboItems": {
        "headers": ["combo_product_key", "item_product_id", "item_product_slug", "variant_id", "variant_sku", "quantity", "sort_order"],
        "rows": [
            ["P010", "", "cotton-t-shirt", "", "TSHIRT-RED-S", 1, 0],
            ["P010", "", "denim-jeans", "", "", 1, 1],
        ],
    },
}


REFERENCE_ROWS = [
    ["Field", "Meaning"],
    ["product_key", "Temporary Excel ID, for example P001. Used to link rows across sheets."],
    ["variant_key", "Temporary Excel variant ID, for example P002-V001."],
    ["type", "Allowed: simple, variable, combo."],
    ["status", "Allowed: draft, published."],
    ["boolean fields", "Use true or false."],
    ["brand_id / brand_slug", "Use either the existing brand UUID or slug."],
    ["category_id / category_slug", "Use either the existing category UUID or slug."],
    ["attribute_id / attribute_slug", "Use either the existing attribute UUID or slug."],
    ["value_id / value_slug", "Use either the existing attribute value UUID or slug."],
    ["media_id / image_url", "Use existing media UUID or a direct image URL."],
    ["simple product", "Requires price and quantity."],
    ["variable product", "Fill ProductVariants and VariantOptions."],
    ["combo product", "Fill ComboItems with existing child products."],
]


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value or "")))
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 14), 45)


def add_validations(ws):
    if ws.title == "Products":
        type_col = ws["E"]
        status_col = ws["F"]
        bool_cols = ["G", "H"]
        type_dv = DataValidation(type="list", formula1='"simple,variable,combo"', allow_blank=False)
        status_dv = DataValidation(type="list", formula1='"draft,published"', allow_blank=False)
        ws.add_data_validation(type_dv)
        ws.add_data_validation(status_dv)
        type_dv.add(f"{type_col[1].coordinate}:E1000")
        status_dv.add(f"{status_col[1].coordinate}:F1000")
        for col in bool_cols:
            dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col}2:{col}1000")
    else:
        for idx, header in enumerate([cell.value for cell in ws[1]], start=1):
            if header in {"is_primary", "is_variant_attr", "is_active"}:
                col = get_column_letter(idx)
                dv = DataValidation(type="list", formula1='"true,false"', allow_blank=True)
                ws.add_data_validation(dv)
                dv.add(f"{col}2:{col}1000")


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)

    for name, spec in SHEETS.items():
        ws = wb.create_sheet(name)
        ws.append(spec["headers"])
        for row in spec["rows"]:
            ws.append(row)
        style_sheet(ws)
        add_validations(ws)

    ref = wb.create_sheet("README")
    for row in REFERENCE_ROWS:
        ref.append(row)
    style_sheet(ref)
    ref.column_dimensions["A"].width = 24
    ref.column_dimensions["B"].width = 90

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
